# sensor_parser.py

from openpyxl import load_workbook
# DATE + TIME PARSING
import re
from datetime import datetime, timedelta, time
# MAIN EXTRACTION
import pandas as pd
import os
from tqdm import tqdm


# ======================
# CONSTANTS
# ======================

GENERATOR_START_ROWS = [2, 36, 70, 104, 138, 172]

TIME_ROW_OFFSET = 6   # ✅ real data starts here
TOTAL_TIME_ROWS = 8   # ✅ 06 → next day 06

# Column mapping (finalized)
COLUMN_MAP = {
    "time": 2,              # B
    "hour_meter": 3,        # C
    "oil_pressure": 4,      # D

    **{f"cyl_{i}_temp": 4 + i for i in range(1, 17)},

    "front_bearing_temp": 21,
    "rear_bearing_temp": 22,
    "oil_temp": 23,
    "coolant_temp": 24,

    "fuel_pressure": 25,
    "diff_fuel_pressure": 26,
    "oil_filter_pressure": 27,
    "air_filter_pressure": 28,
    "crankcase_pressure": 29,

    "kw": 30,
    "kva": 31,
    "cos_phi": 32,
    "amp_r": 33,
    "amp_s": 34,
    "amp_t": 35,
    "load_percent": 36,
}

MONTH_MAP = {
    "Januari": "January",
    "Februari": "February",
    "Maret": "March",
    "April": "April",
    "Mei": "May",
    "Juni": "June",
    "Juli": "July",
    "Agustus": "August",
    "September": "September",
    "Oktober": "October",
    "November": "November",
    "Desember": "December"
}

LOCATION_BLOCKS = {
    "CHPP": (6, 11),
    "MCC": (12, 17),
    "WATER": (18, 23),
    "MIA": (24, 30),
}

# ======================
# HELPER FUNCTIONS
# ======================

def extract_year_month_from_filename(file_name):
    year_match = re.search(r"(20\d{2})", file_name)
    if not year_match:
        raise ValueError(f"❌ Cannot extract year from filename: {file_name}")

    year = int(year_match.group(1))

    for indo, eng in MONTH_MAP.items():
        if indo.lower() in file_name.lower():
            month = datetime.strptime(eng, "%B").month
            return year, month

    raise ValueError(f"❌ Cannot extract month from filename: {file_name}")

def extract_date(sheet_name, file_year, file_month):
    sheet_name = sheet_name.replace("Tgl", "").strip()

    # 🔥 FIXED — allow 01, 02, etc
    match = re.search(r'\b(0?[1-9]|[12][0-9]|3[01])\b', sheet_name)

    if not match:
        raise ValueError(f"❌ Cannot extract DAY from sheet: {sheet_name}")

    day = int(match.group(1))

    return datetime(file_year, file_month, day)

def clean_time(val):
    if isinstance(val, str):
        val = val.replace("'", ":").strip()

        if val == "24:00":
            return 24, 0  # keep as signal

        try:
            t = datetime.strptime(val, "%H:%M").time()
            return t.hour, t.minute
        except ValueError:
            return None

    elif isinstance(val, time):
        return val.hour, val.minute

    return None


def combine_datetime(date, hour, minute, day_offset):
    return datetime.combine(date.date(), time(hour % 24, minute)) + timedelta(days=day_offset)

# GENERATOR METADATA
def extract_generator_name(sheet, start_row):
    cell_val = sheet.cell(start_row + 1, 6).value  # F column

    if not cell_val:
        return None

    match = re.search(r'PSPWGE-\d+', str(cell_val))
    return match.group(0) if match else None


def get_total_ge(sheet, start_row):
    return sheet.cell(start_row + 32, 38).value  # AL column


def is_generator_running(total_ge):
    try:
        total_ge = float(total_ge)
        return total_ge > 0
    except (TypeError, ValueError):
        return False
    
# ROW LOGIC to determine if row is empty or not and to detetect if machine is running
def is_row_empty(sheet, row):
    # Check FULL data range (D → AP)
    for col in range(4, 37):
        val = sheet.cell(row, col).value
        if val not in [None, ""]:
            return False
    return True


def determine_state(is_running_today, row_empty):
    if not is_running_today:
        return "OFF"

    return "RUNNING"

def extract_row(sheet, row, timestamp, generator_id, is_running_today):
    row_empty = is_row_empty(sheet, row)

    state = determine_state(is_running_today, row_empty)

    record = {
        "timestamp": timestamp,
        "generator_id": generator_id,
        "operating_state": state,
        "is_missing": row_empty,
    }

    # 🔥 define fields that should survive even if row is "missing"
    ALWAYS_KEEP_FIELDS = ["hour_meter"]  # add more if needed later

    # Fill values
    for key, col in COLUMN_MAP.items():
        if key == "time":
            continue

        val = sheet.cell(row, col).value

        # 🔥 CASE 1: machine OFF → everything = 0
        if not is_running_today:
            record[key] = 0

        # 🔥 CASE 2: row missing BUT field is important → keep it
        elif row_empty and key in ALWAYS_KEEP_FIELDS:
            record[key] = val if val is not None else 0

        # 🔥 CASE 3: row missing → zero out normal sensor fields
        elif row_empty:
            record[key] = 0

        # 🔥 CASE 4: normal case
        else:
            record[key] = val if val is not None else 0

    return record

# FINAL DF STRUCTURE
def build_schema():
    return [
        "timestamp",
        "generator_id",
        "operating_state",
        "is_missing",

        "hour_meter",
        "oil_pressure",

        *[f"cyl_{i}_temp" for i in range(1, 17)],

        "front_bearing_temp",
        "rear_bearing_temp",
        "oil_temp",
        "coolant_temp",

        "fuel_pressure",
        "diff_fuel_pressure",
        "oil_filter_pressure",
        "air_filter_pressure",
        "crankcase_pressure",

        "kw",
        "kva",
        "cos_phi",
        "amp_r",
        "amp_s",
        "amp_t",
        "load_percent",
    ]

def build_daily_summary_schema():
    return [
        "date",
        "generator_id",

        "location",  # CHPP / MCC / WATER / MIA

        "kwh_prev_month",
        "kwh_daily",
        "kwh_cumulative",

        "fuel_fm_total",   # shared (same for all 6 gens)
        "fuel_ge_total",   # per generator
    ]

def extract_daily_summary(sheet, start_row, date, generator_id):

    records = []

    fuel_fm = sheet.cell(start_row + 32, 37).value  # AK
    fuel_ge = sheet.cell(start_row + 32, 38).value  # AL

    for location, (start_offset, end_offset) in LOCATION_BLOCKS.items():

        # ✅ take ONLY the first row of the block
        row = start_row + start_offset

        record = {
            "date": date,
            "generator_id": generator_id,
            "location": location,

            "kwh_prev_month": sheet.cell(row, 40).value,
            "kwh_daily": sheet.cell(row, 41).value,
            "kwh_cumulative": sheet.cell(row, 42).value,

            "fuel_fm_total": fuel_fm,
            "fuel_ge_total": fuel_ge,
        }

        records.append(record)

    return records

def ingest_folder(base_path, max_sheets=None):
    all_main_dfs = []
    all_summary_dfs = []

    # 🔥 collect all file paths first
    file_paths = []

    for year_folder in os.listdir(base_path):
        year_path = os.path.join(base_path, year_folder)

        if not os.path.isdir(year_path):
            continue

        for file_name in os.listdir(year_path):
            if file_name.endswith(".xlsx"):
                file_paths.append(os.path.join(year_path, file_name))

    print(f"Total files to process: {len(file_paths)}")

    # 🔥 progress bar
    for file_path in tqdm(file_paths, desc="Processing files"):

        file_name = os.path.basename(file_path)

        try:
            df, summary_df = extract_dataframe(
                file_path,
                max_sheets=max_sheets,
            )

            df["source_file"] = file_name
            summary_df["source_file"] = file_name

            all_main_dfs.append(df)
            all_summary_dfs.append(summary_df)

        except Exception as e:
            print(f"\n❌ Error in {file_name}: {e}")

    if not all_main_dfs:
        raise ValueError("No valid sensor files processed.")

    if not all_summary_dfs:
        raise ValueError("No valid summary files processed.")
    
    final_df = pd.concat(all_main_dfs, ignore_index=True)
    final_summary_df = pd.concat(all_summary_dfs, ignore_index=True)

    final_df["timestamp"] = pd.to_datetime(final_df["timestamp"])
    final_summary_df["date"] = pd.to_datetime(final_summary_df["date"])

    # =========================
    # 🔥 MAIN SENSOR DATA SORT
    # =========================
    final_df = final_df.sort_values(
        ["generator_id", "timestamp"]
    ).reset_index(drop=True)

    # =========================
    # 🔥 SUMMARY DATA FIX
    # =========================

    # 1. preserve ingestion order
    final_summary_df["ingest_order"] = final_summary_df.index

    # 2. deduplicate BEFORE sorting (keep first = your "gold")
    final_summary_df = final_summary_df.sort_values(
        ["generator_id", "date", "location", "ingest_order"]
    )

    final_summary_df = final_summary_df.drop_duplicates(
        subset=["generator_id", "date", "location"],
        keep="first"
    )

    # 3. final clean sort
    final_summary_df = final_summary_df.sort_values(
        ["generator_id", "date"]
    ).reset_index(drop=True)

    # optional cleanup
    final_summary_df = final_summary_df.drop(columns=["ingest_order"])

    final_summary_df = final_summary_df[
    final_summary_df["generator_id"].notna() &
    final_summary_df["date"].notna()
]

    final_df = final_df[
    final_df["generator_id"].notna() &
    final_df["timestamp"].notna()
]


    return final_df, final_summary_df

def save_with_progress(df, file_name, chunk_size=50000):
    total_rows = len(df)

    with open(file_name, "w", encoding="utf-8") as f:
        # write header first
        df.head(0).to_csv(f, index=False)

        # write in chunks
        for i in tqdm(range(0, total_rows, chunk_size), desc=f"Saving {file_name}"):
            chunk = df.iloc[i:i+chunk_size]
            chunk.to_csv(f, index=False, header=False, lineterminator="\n")

# ======================
# MAIN PARSER
# ======================

# MAIN EXTRACTION
def extract_dataframe(file_path, max_sheets=None):
    
    file_name = os.path.basename(file_path)
    file_year, file_month = extract_year_month_from_filename(file_name)

    wb = load_workbook(file_path, data_only=True)

    if max_sheets is not None:
        sheets = wb.worksheets[:max_sheets]
    else:
        sheets = wb.worksheets

    all_rows = []
    all_summary_rows = []   # ✅ NEW

    for sheet_index, sheet in enumerate(sheets):  # 🔥 NEW: track sheet position
        date = extract_date(sheet.title, file_year, file_month)
        if not date:
            continue

        for start_row in GENERATOR_START_ROWS:  # repeats 6x for each generator

            generator_id = extract_generator_name(sheet, start_row)
            total_ge = get_total_ge(sheet, start_row)
            is_running_today = is_generator_running(total_ge)

            # ✅ NEW — extract daily summary ONCE per generator
            summary_records = extract_daily_summary(
                sheet,
                start_row,
                date,
                generator_id
            )
            all_summary_rows.extend(summary_records)

            # existing 3-hour loop (UNCHANGED)
            ROW_STEP = 3

            previous_hour = None
            day_offset = 0

            for i in range(TOTAL_TIME_ROWS):

                # # 🔥 NEW — skip first 06:00 row for all sheets except the first
                # if sheet_index > 0 and i == 0:
                #     continue

                row = start_row + TIME_ROW_OFFSET + (i * ROW_STEP)

                time_val = sheet.cell(row, COLUMN_MAP["time"]).value
                cleaned = clean_time(time_val)

                if cleaned is None:
                    continue

                current_hour, current_minute = cleaned

                # 🔥 CASE 1: 24:00 → force next day
                if current_hour == 24:
                    day_offset += 1

                # 🔥 CASE 2: rollover (but NOT right after 24)
                elif previous_hour is not None and previous_hour != 24 and current_hour < previous_hour:
                    day_offset += 1

                timestamp = combine_datetime(date, current_hour, current_minute, day_offset)

                record = extract_row(
                    sheet,
                    row,
                    timestamp,
                    generator_id,
                    is_running_today
                )

                all_rows.append(record)

                previous_hour = current_hour

    # ✅ build main df
    df = pd.DataFrame(all_rows)
    schema = build_schema()
    for col in schema:
        if col not in df.columns:
            df[col] = 0

    df = df[schema]
    df = df.sort_values(["generator_id", "timestamp"])
    df = df.drop_duplicates(subset=["generator_id", "timestamp"], keep="first")

    # ✅ build summary df (NEW)
    summary_df = pd.DataFrame(all_summary_rows)
    summary_schema = build_daily_summary_schema()

    for col in summary_schema:
        if col not in summary_df.columns:
            summary_df[col] = None

    summary_df = summary_df[summary_schema]

    return df, summary_df   # ✅ CHANGED (2 outputs)


# base_path = "/content/01. Data Daily Genset"
# df, summary_df = ingest_folder(base_path)

# # 🔥 save both datasets with progress
# save_with_progress(df, "sensor_data.csv")
# save_with_progress(summary_df, "daily_summary.csv")

# # 🔥 download
# from google.colab import files
# files.download("sensor_data.csv")
# files.download("daily_summary.csv")